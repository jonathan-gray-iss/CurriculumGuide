#! /usr/local/bin/python3
import csv
import pickle
from helper import *
import copy
import openpyxl

def fix_terms(terms):
    if terms == "F":
        return "Fall Semester"
    elif terms == "S":
        return "Spring Semester"
    elif terms == "FS":
        return "Fall and Spring Semesters"
    elif terms == "Y":
        return "Year"
        
def fix_credits(credits):
    if credits == "0.5" or credits == 0.5:
        return "0.5 Credits"        
    elif credits == "1.0" or credits == 1.0:
        return "1 Credit"
    elif credits == "1" or credits == 1:
        return "1 Credit"     
    else:
        return credits   

def ind_study(string):
    if "IS" in string:
        try:
            s = string[0:string.index("-IS")] 
            return s, 1
        except:
            s = string[0:string.index("- IS")]  
            return s, 1
        
        try: 
            s = string[0:string.index("IS")] 
        except:
            pass
    else:
        return string, 0            
    
def make_entry(course_info):


    
    title = course_info [0]
    terms = course_info[1]
    credits = fix_credits(course_info[2])
    teacher = course_info[3]
    desc = "\\emph{"+course_info[4]+"}"

    terms = fix_terms(terms)
    

    
    
    
    

    if course_info[4] != "OLD":

    
        p1 = "\\noindent\\textbf{"
        p2 = str(title)
        p3 = "} \\hfill "
        p6 = str(teacher)    
        p4 = str(terms)
        p5 = " - "
        p7 = str(credits)

        p = p1 + p2 + p3 + p6 
        p += "\n\n" + "\\noindent "
        p += p4 + p5 + p7 
        p += "\n\n" + "\\vspace{1mm}"
        p += desc 
        p += "\\\\" + "\n\n"
    
    elif course_info[4] == "OLD":

        p1 = "\\noindent\\textbf{"
        p2 = str(title)
        p3 = "} "  
        p5 = " - "
        p7 = str(credits)  
        
        p = p1 + p2 + p3  
        p +=  p5 + p7 
        p += "\n\n" + "\\vspace{3mm}"            
    
    
    return p
                                         
        
def csv_processor():


    return records

#####
#####
##### Set Grad Year (as str)
#####
#####


csvFile = "Courses_testcases.csv"
xlsFile = "courses2324.xlsx"


write_now = False    
wb = openpyxl.load_workbook(xlsFile, data_only=True)  
print(wb.sheetnames)


for name in wb.sheetnames:
    
    wdir = "./Chapters/"+name
    wfile = "/Sect2.tex"
    print(wdir+wfile)
    
    to_write = ""
    
    ws = wb[name]

    print("\n"+"#"*10 +" " + name +" "+ "#"*10+"\n")
   
    first_new = True
    first_old = True    



    for row in range(1,ws.max_row):
        bool1 = (ws.cell(row,1).value is not None)
        bool2 = (ws.cell(row,2).value != "Course Title")
        if bool1 and bool2:
            
            cnum = ws.cell(row,1).value
            cname = ws.cell(row,2).value   
            cdesc = ws.cell(row,3).value                
            ccred = ws.cell(row,4).value           
            cdept = ws.cell(row,5).value
            cterms = ws.cell(row,6).value
            cteach = ws.cell(row,7).value            


            if "&" in cdesc:
                cdesc = cdesc.split("&")             
                cdesc = cdesc[0] + "\\&" + cdesc[1]

            if "&" in cname:

                cname = cname.split("&")             
                cname = cname[0] + "\\&" + cname[1]

            bundle = [cname, cterms,ccred,cteach,cdesc]
            
            if name == "IS":
                print("\\noindent\\textbf{"+cname+"}"+"\n")
            
            if first_new == True:
                to_write += "\\subsection{Current Courses}"+"\n"
                first_new = False
                
          
            if "OLD" in cdesc:
                if first_old == True:

                    to_write += "\\subsection{Past Courses}"+"\n"
                    first_old = False
            
            out = make_entry(bundle)
            to_write += out
    
    
    
    if write_now == True:
        with open(wdir+wfile, "w") as f:
            f.write("\section{Course Descriptions}\n\n")
            f.write(to_write)


        


    




#for line in records:
#    print line
#
#    title_query = ind_study(line[0])
#
#    if title_query[0][-3:] == "- F":
#        title = title_query[0][:-3]
#
#    elif title_query[0][-3:] == "- S":
#        title = title_query[0][:-3]  
#    else:
#        title = title_query[0]      
#    
#    desc = line[1]
#    credits = line[2]
#    dept = line[3]
#    if title_query[1] == 1:
#        dept = "Individualized Learning"   
#    new_records.append([title,desc,credits,dept])    
    



    
 
        

    











 

 
